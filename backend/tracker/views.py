import csv
import json
from datetime import date, timedelta

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils.safestring import mark_safe
from django.db.models import Count, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST

from .forms import DailyEntryForm, EntryItemFormSet, ExtraEntryItemFormSet, PlanUpdateLineFormSet, AEDailyUpdateForm, TASK_TYPE_CHOICES
from .models import DailyEntry, EntryItem, Member, AEDailyUpdate


# ── helpers ──────────────────────────────────────────────────────────────────

def _plan_for_member_date(member_pk, entry_d):
    """Return (plan DailyEntry or None, ordered plan EntryItems) for updates."""
    if not member_pk:
        return None, []
    plan = (
        DailyEntry.objects.filter(
            kind='plan',
            member_id=member_pk,
            entry_date=entry_d,
        )
        .prefetch_related('items')
        .first()
    )
    if not plan:
        return None, []
    return plan, list(plan.items.all().order_by('id'))


def _plan_items_ordered_from_formset(plan_line_formset):
    ids = []
    for f in plan_line_formset.forms:
        raw = f.data.get(f.add_prefix('plan_item_id'))
        if raw is None or raw == '':
            raw = f.initial.get('plan_item_id')
        if raw is None or raw == '':
            continue
        try:
            ids.append(int(raw))
        except (TypeError, ValueError):
            continue
    if not ids:
        return []
    by_id = {i.pk: i for i in EntryItem.objects.filter(pk__in=ids)}
    return [by_id[i] for i in ids if i in by_id]


def _plan_line_rows(plan_line_formset, plan_items):
    rows = []
    for i, pit in enumerate(plan_items):
        if i < len(plan_line_formset.forms):
            rows.append({'item': pit, 'form': plan_line_formset.forms[i]})
    return rows


def _monday(d):
    return d - timedelta(days=d.weekday())

def _parse_date(s, default=None):
    if not s:
        return default
    try:
        from datetime import datetime
        return datetime.strptime(s, '%Y-%m-%d').date()
    except Exception:
        return default


# ── public API ───────────────────────────────────────────────────────────────

def api_health(request):
    from datetime import datetime
    return JsonResponse({'ok': True, 'time': datetime.utcnow().isoformat() + 'Z'})


@csrf_exempt
def api_intake_slack(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Method not allowed'}, status=405)

    expected = getattr(settings, 'INTAKE_TOKEN', '') or ''
    if expected:
        token = (request.POST.get('token') or request.headers.get('x-intake-token') or '')
        if token != expected:
            return JsonResponse({'ok': False, 'error': 'Unauthorized'}, status=401)

    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except Exception:
        return JsonResponse({'ok': False, 'error': 'Invalid JSON'}, status=400)

    member_name = (payload.get('member') or '').strip()
    kind = (payload.get('kind') or '').strip().lower()
    if not member_name or kind not in ('plan', 'update'):
        return JsonResponse({'ok': False, 'error': 'Missing member/kind'}, status=400)

    entry_date = _parse_date(payload.get('date'), default=date.today())
    status = payload.get('status', 'open')
    if status not in ('open', 'in_progress', 'blocked', 'closed'):
        status = 'open'

    member, _ = Member.objects.get_or_create(display_name=member_name)
    entry = DailyEntry.objects.create(
        entry_date=entry_date, kind=kind, status=status,
        member=member, raw_text=payload.get('rawText') or None, source='slack',
    )
    count = 0
    for it in payload.get('items') or []:
        task_type = (it.get('taskType') or it.get('task_type') or '').strip()
        if not task_type:
            continue
        EntryItem.objects.create(
            entry=entry, task_type=task_type,
            question_type=it.get('questionType') or None,
            customer=it.get('customer') or None,
            count=it.get('count') or None,
            notes=it.get('notes') or None,
        )
        count += 1

    response = {'ok': True, 'entryId': entry.id, 'itemsCreated': count}
    if kind == 'plan':
        from .jira_client import create_item_issue

        jira_keys, jira_errs = [], []
        entry.refresh_from_db()
        for it in entry.items.all():
            jr = create_item_issue(entry, it)
            if jr.get('ok'):
                it.jira_issue_key = jr['key']
                it.jira_issue_url = jr['url']
                it.save(update_fields=['jira_issue_key', 'jira_issue_url'])
                jira_keys.append(jr['key'])
            elif jr.get('error'):
                jira_errs.append(f"{it.task_type}: {jr['error']}")
        if jira_keys:
            response['jiraKeys'] = jira_keys
        if jira_errs:
            response['jiraError'] = '; '.join(jira_errs)

    from .slack_notify import notify_entry_slack_thread
    notify_entry_slack_thread(entry)
    return JsonResponse(response)


# ── dashboard ─────────────────────────────────────────────────────────────────

@login_required
def dashboard(request):
    today = date.today()
    week_start = _monday(today)
    week_end = week_start + timedelta(days=6)

    from_d = _parse_date(request.GET.get('from'), default=week_start)
    to_d   = _parse_date(request.GET.get('to'),   default=week_end)
    sel_member = (request.GET.get('member') or '').strip()
    sel_kind   = (request.GET.get('kind') or '').strip().lower()
    sel_status = (request.GET.get('status') or '').strip().lower()
    q          = (request.GET.get('q') or '').strip()

    qs = DailyEntry.objects.select_related('member').prefetch_related('items').filter(
        entry_date__range=(from_d, to_d)
    )
    if sel_member:
        qs = qs.filter(member__display_name=sel_member)
    if sel_kind in ('plan', 'update'):
        qs = qs.filter(kind=sel_kind)
    if sel_status in ('open', 'in_progress', 'blocked', 'closed'):
        qs = qs.filter(status=sel_status)
    if q:
        from django.db.models import Q
        qs = qs.filter(
            Q(raw_text__icontains=q) |
            Q(items__notes__icontains=q) |
            Q(items__task_type__icontains=q) |
            Q(items__customer__icontains=q)
        ).distinct()

    flat = []
    for e in qs[:500]:
        items = list(e.items.all())
        base = dict(
            id=e.id,
            date=e.entry_date.isoformat(),
            kind=e.kind,
            status=e.status,
            member=e.member.display_name,
            jira_key=e.jira_issue_key or '',
            jira_url=e.jira_issue_url or '',
        )
        if not items:
            flat.append({
                **base,
                'item_id': None,
                'task_type': '',
                'question_type': '',
                'customer': '',
                'count': '',
                'notes': e.raw_text or '',
                'status': e.status,
                'jira_key': e.jira_issue_key or '',
                'jira_url': e.jira_issue_url or '',
            })
        else:
            for it in items:
                jk = it.jira_issue_key or e.jira_issue_key or ''
                ju = it.jira_issue_url or e.jira_issue_url or ''
                flat.append({
                    **base,
                    'item_id': it.id,
                    'task_type': it.task_type,
                    'question_type': it.question_type or '',
                    'customer': it.customer or '',
                    'count': it.count or '',
                    'notes': it.notes or e.raw_text or '',
                    'status': it.status,
                    'jira_key': jk,
                    'jira_url': ju,
                })

    kpis = {
        'members':       len({r['member'] for r in flat}),
        'items':         len([r for r in flat if r['task_type']]),
        'plans':         len({r['id'] for r in flat if r['kind'] == 'plan'}),
        'updates':       len({r['id'] for r in flat if r['kind'] == 'update'}),
        'today_plans':   DailyEntry.objects.filter(kind='plan',   entry_date=today).count(),
        'today_updates': DailyEntry.objects.filter(kind='update', entry_date=today).count(),
        'open_items':    EntryItem.objects.filter(entry__entry_date__range=(from_d, to_d), status='open').count(),
    }

    # chart: plan + update entries per day (not volume)
    by_date = {}
    for r in flat:
        by_date.setdefault(r['date'], {'plans': set(), 'updates': set()})
        if r['kind'] == 'plan':
            by_date[r['date']]['plans'].add(r['id'])
        else:
            by_date[r['date']]['updates'].add(r['id'])
    points, cur = [], from_d
    while cur <= to_d:
        key = cur.isoformat()
        d = by_date.get(key, {})
        points.append({
            'date':    cur.strftime('%b %d'),
            'plans':   len(d.get('plans', set())),
            'updates': len(d.get('updates', set())),
        })
        cur += timedelta(days=1)

    # analysis data ─ by member
    member_stats = {}
    for r in flat:
        m = r['member']
        s = member_stats.setdefault(m, {'member': m, 'plans': set(), 'updates': set(),
                                        'items': 0, 'open': 0, 'in_progress': 0, 'closed': 0, 'blocked': 0})
        if r['kind'] == 'plan':   s['plans'].add(r['id'])
        if r['kind'] == 'update': s['updates'].add(r['id'])
        if r['task_type']:
            s['items'] += 1
            st = r['status']
            if st in ('open', 'in_progress', 'closed', 'blocked'):
                s[st] += 1
    analysis_by_member = sorted(
        [{'member': v['member'], 'plans': len(v['plans']), 'updates': len(v['updates']),
          'items': v['items'], 'open': v['open'], 'in_progress': v['in_progress'],
          'closed': v['closed'], 'blocked': v['blocked']}
         for v in member_stats.values()],
        key=lambda x: -x['items'],
    )

    # analysis ─ by work type
    work_type_stats = {}
    for r in flat:
        tt = r['task_type'] or '(no type)'
        s = work_type_stats.setdefault(tt, {'task_type': tt, 'items': 0, 'open': 0,
                                            'in_progress': 0, 'closed': 0, 'blocked': 0})
        s['items'] += 1
        st = r['status']
        if st in ('open', 'in_progress', 'closed', 'blocked'):
            s[st] += 1
    analysis_by_worktype = sorted(work_type_stats.values(), key=lambda x: -x['items'])

    # analysis ─ open items detail
    open_rows = [r for r in flat if r['status'] == 'open' and r['task_type']]

    members = Member.objects.filter(is_active=True).order_by('display_name')

    return render(request, 'tracker/dashboard.html', {
        'from': from_d.isoformat(), 'to': to_d.isoformat(),
        'sel_member': sel_member, 'sel_kind': sel_kind,
        'sel_status': sel_status, 'q': q,
        'kpis': kpis,
        'kpis_json':            mark_safe(json.dumps(kpis)),
        'points_json':          mark_safe(json.dumps(points)),
        'members': members,
        'rows': flat[:250],
        'today': today.isoformat(),
        'analysis_by_member':      analysis_by_member,
        'analysis_by_worktype':    analysis_by_worktype,
        'analysis_by_member_json': mark_safe(json.dumps(analysis_by_member)),
        'analysis_by_worktype_json': mark_safe(json.dumps(analysis_by_worktype)),
        'open_rows':            open_rows[:200],
    })


# ── create entry ──────────────────────────────────────────────────────────────

def _save_plan_with_jira(request, entry):
    """Create one Jira issue per plan task; leave entry-level Jira fields empty."""
    from .jira_client import create_item_issue

    entry.jira_issue_key = None
    entry.jira_issue_url = None
    entry.save(update_fields=['jira_issue_key', 'jira_issue_url'])
    keys, errs = [], []
    for it in entry.items.all():
        jr = create_item_issue(entry, it)
        if jr.get('ok'):
            it.jira_issue_key = jr['key']
            it.jira_issue_url = jr['url']
            it.save(update_fields=['jira_issue_key', 'jira_issue_url'])
            keys.append((jr['key'], jr['url']))
        elif jr.get('error'):
            errs.append(f"{it.task_type}: {jr['error']}")
    if keys:
        links = ' &nbsp;·&nbsp; '.join(
            f'<a href="{u}" target="_blank" rel="noopener noreferrer" class="link-jira">{k}</a>'
            for k, u in keys
        )
        messages.success(
            request,
            mark_safe(f'Plan saved. Jira ({len(keys)} ticket{"s" if len(keys) != 1 else ""}): {links}'),
        )
    else:
        messages.success(request, 'Plan saved.')
    for err in errs:
        messages.warning(request, err)


@login_required
def entry_create(request, kind):
    today = date.today()
    ctx_common = {
        'kind': kind,
        'kind_label': 'Plan' if kind == 'plan' else 'Update',
    }

    if kind == 'update':
        if request.method == 'POST':
            form = DailyEntryForm(request.POST)
            plan_line_formset = PlanUpdateLineFormSet(request.POST, prefix='planlines')
            extra_formset = EntryItemFormSet(request.POST, prefix='extra')
            fv = form.is_valid()
            plv = plan_line_formset.is_valid()
            exv = extra_formset.is_valid()
            valid = fv and plv and exv
            if valid:
                member = form.cleaned_data['member']
                entry_date = form.cleaned_data['entry_date']
                plan, plan_items = _plan_for_member_date(member.pk, entry_date)
                allowed = {it.id for it in plan_items}
                n_forms = plan_line_formset.total_form_count()
                if n_forms > 0 and not allowed:
                    messages.error(
                        request,
                        'No plan exists for this member and date — remove stale plan rows or pick another date.',
                    )
                    valid = False
                elif n_forms > 0:
                    for row in plan_line_formset.cleaned_data:
                        if not row:
                            continue
                        if row['plan_item_id'] not in allowed:
                            messages.error(
                                request,
                                'Plan tasks do not match this member and date. Reload the form.',
                            )
                            valid = False
                            break
            if valid:
                from .jira_client import create_item_issue, sync_item_jira_status

                entry = form.save(commit=False)
                entry.kind = 'update'
                entry.source = 'web'
                entry.jira_issue_key = None
                entry.jira_issue_url = None
                entry.save()
                jira_msgs = []
                for row in plan_line_formset.cleaned_data:
                    if not row:
                        continue
                    plan_item = EntryItem.objects.get(pk=row['plan_item_id'])
                    notes = (row.get('notes') or '').strip() or None
                    due_at = row.get('due_at')
                    sync_r = sync_item_jira_status(plan_item, row['status'], notes, due_at=due_at)
                    if not sync_r.get('ok') and not sync_r.get('skipped'):
                        jira_msgs.append(
                            f"{plan_item.task_type}: {sync_r.get('error', 'Jira')}"
                        )
                    plan_item.status = row['status']
                    u_fields = ['status']
                    if row.get('count') is not None:
                        plan_item.count = row['count']
                        u_fields.append('count')
                    if due_at:
                        plan_item.due_at = due_at
                        u_fields.append('due_at')
                    plan_item.save(update_fields=u_fields)
                    EntryItem.objects.create(
                        entry=entry,
                        plan_item=plan_item,
                        task_type=plan_item.task_type,
                        question_type=plan_item.question_type,
                        customer=plan_item.customer,
                        count=row.get('count'),
                        notes=notes,
                        status=row['status'],
                        jira_issue_key=plan_item.jira_issue_key,
                        jira_issue_url=plan_item.jira_issue_url,
                        due_at=due_at,
                    )
                for f in extra_formset.cleaned_data:
                    if not f or f.get('DELETE'):
                        continue
                    task_type = (f.get('task_type') or '').strip()
                    if not task_type:
                        continue
                    extra_status = f.get('status') or 'open'
                    if extra_status not in ('open', 'in_progress', 'blocked', 'closed'):
                        extra_status = 'open'
                    it = EntryItem.objects.create(
                        entry=entry,
                        task_type=task_type,
                        question_type=(f.get('question_type') or '').strip() or None,
                        customer=(f.get('customer') or '').strip() or None,
                        count=f.get('count') or None,
                        notes=(f.get('notes') or '').strip() or None,
                        due_at=f.get('due_at') or None,
                        status=extra_status,
                    )
                    jr = create_item_issue(entry, it)
                    if jr.get('ok'):
                        it.jira_issue_key = jr['key']
                        it.jira_issue_url = jr['url']
                        it.save(update_fields=['jira_issue_key', 'jira_issue_url'])
                    elif jr.get('error'):
                        jira_msgs.append(f"{task_type}: {jr['error']}")
                messages.success(request, 'Update saved.')
                for m in jira_msgs:
                    messages.warning(request, m)
                entry.refresh_from_db()
                from .slack_notify import notify_entry_slack_thread
                notify_entry_slack_thread(entry)
                return redirect('dashboard')
            plan_items = _plan_items_ordered_from_formset(plan_line_formset)
            plan, _ = _plan_for_member_date(
                int(request.POST['member']) if request.POST.get('member', '').isdigit() else None,
                _parse_date(request.POST.get('entry_date'), today),
            )
            plan_line_rows = _plan_line_rows(plan_line_formset, plan_items)
        else:
            init_date = _parse_date(request.GET.get('date'), default=today)
            init = {'entry_date': init_date}
            mq = request.GET.get('member', '').strip()
            if mq.isdigit():
                init['member'] = int(mq)
            form = DailyEntryForm(initial=init)
            member_pk = init.get('member')
            plan, plan_items = _plan_for_member_date(member_pk, init_date)
            initial_lines = [
                {
                    'plan_item_id': x.id,
                    'status': x.status,
                    'count': x.count,
                    'notes': '',
                    'due_at': x.due_at,
                }
                for x in plan_items
            ]
            plan_line_formset = PlanUpdateLineFormSet(
                initial=initial_lines, prefix='planlines'
            )
            extra_formset = ExtraEntryItemFormSet(prefix='extra')
            plan_line_rows = _plan_line_rows(plan_line_formset, plan_items)

        return render(
            request,
            'tracker/entry_form.html',
            {
                **ctx_common,
                'form': form,
                'formset': None,
                'plan_entry': plan if kind == 'update' else None,
                'plan_items': plan_items,
                'plan_line_formset': plan_line_formset,
                'plan_line_rows': plan_line_rows,
                'extra_formset': extra_formset,
                'task_formset_prefix': 'extra',
                'task_type_choices': TASK_TYPE_CHOICES,
                'task_type_choices_json': json.dumps([
                    {'value': v, 'label': l} for v, l in TASK_TYPE_CHOICES if v
                ]),
            },
        )

    # ── plan ───────────────────────────────────────────────────────────────
    if request.method == 'POST':
        form = DailyEntryForm(request.POST)
        formset = EntryItemFormSet(request.POST, prefix='items')
        if form.is_valid() and formset.is_valid():
            entry = form.save(commit=False)
            entry.kind = 'plan'
            entry.source = 'web'
            entry.jira_issue_key = None
            entry.jira_issue_url = None
            entry.save()
            for f in formset.cleaned_data:
                if not f or f.get('DELETE'):
                    continue
                task_type = (f.get('task_type') or '').strip()
                if not task_type:
                    continue
                EntryItem.objects.create(
                    entry=entry,
                    task_type=task_type,
                    question_type=(f.get('question_type') or '').strip() or None,
                    customer=(f.get('customer') or '').strip() or None,
                    count=f.get('count') or None,
                    notes=(f.get('notes') or '').strip() or None,
                    due_at=f.get('due_at') or None,
                    status='open',
                )
            _save_plan_with_jira(request, entry)
            entry.refresh_from_db()
            from .slack_notify import notify_entry_slack_thread
            notify_entry_slack_thread(entry)
            return redirect('dashboard')
    else:
        form = DailyEntryForm(initial={'entry_date': today})
        formset = EntryItemFormSet(prefix='items')

    return render(
        request,
        'tracker/entry_form.html',
        {
            **ctx_common,
            'form': form,
            'formset': formset,
            'plan_entry': None,
            'plan_items': [],
            'plan_line_formset': None,
            'plan_line_rows': [],
            'extra_formset': None,
            'task_formset_prefix': 'items',
            'task_type_choices': TASK_TYPE_CHOICES,
            'task_type_choices_json': json.dumps([
                {'value': v, 'label': lb} for v, lb in TASK_TYPE_CHOICES if v
            ]),
        },
    )


# ── update status (AJAX/form POST) ────────────────────────────────────────────

@login_required
@require_POST
def update_status(request):
    try:
        entry_id = int(request.POST.get('entry_id', ''))
    except (TypeError, ValueError):
        return redirect(request.META.get('HTTP_REFERER', '/'))

    new_status = request.POST.get('status', '')
    if new_status not in ('open', 'in_progress', 'blocked', 'closed'):
        return redirect(request.META.get('HTTP_REFERER', '/'))

    from .jira_client import transition_issue

    comment = (request.POST.get('comment') or '').strip() or None
    due_raw = (request.POST.get('due_at') or '').strip() or None
    due_at = None
    if due_raw:
        try:
            from datetime import datetime
            due_at = datetime.strptime(due_raw, '%Y-%m-%d').date()
        except Exception:
            due_at = None

    item_raw = (request.POST.get('entry_item_id') or '').strip()
    if item_raw.isdigit():
        item = get_object_or_404(
            EntryItem.objects.select_related('plan_item', 'entry'),
            pk=int(item_raw),
        )
        if item.entry_id != entry_id:
            return redirect(request.META.get('HTTP_REFERER', '/'))
        item.status = new_status
        item.save(update_fields=['status'])
        if item.plan_item_id:
            EntryItem.objects.filter(pk=item.plan_item_id).update(status=new_status)
            jkey = item.jira_issue_key or (
                item.plan_item.jira_issue_key if item.plan_item else ''
            )
        else:
            EntryItem.objects.filter(plan_item_id=item.pk).update(status=new_status)
            jkey = item.jira_issue_key or ''
        if jkey:
            tr = transition_issue(jkey, new_status, comment=comment, due_at=due_at)
            if not tr.get('ok'):
                messages.warning(request, tr.get('error', 'Jira update failed'))
        ref_status = new_status
    else:
        entry = get_object_or_404(DailyEntry, id=entry_id)
        entry.status = new_status
        entry.save(update_fields=['status'])
        seen_keys = set()
        for it in EntryItem.objects.filter(entry=entry).select_related('plan_item'):
            it.status = new_status
            it.save(update_fields=['status'])
            k = it.jira_issue_key or ''
            if not k and it.plan_item_id and it.plan_item:
                k = it.plan_item.jira_issue_key or ''
            if k and k not in seen_keys:
                seen_keys.add(k)
                tr = transition_issue(k, new_status, comment=comment, due_at=due_at)
                if not tr.get('ok'):
                    messages.warning(request, tr.get('error', 'Jira update failed'))
        if entry.jira_issue_key and entry.jira_issue_key not in seen_keys:
            tr = transition_issue(entry.jira_issue_key, new_status, comment=comment, due_at=due_at)
            if not tr.get('ok'):
                messages.warning(request, tr.get('error', 'Jira update failed'))
        ref_status = entry.status

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'ok': True, 'status': ref_status})
    return redirect(request.META.get('HTTP_REFERER', '/'))


# ── extra tasks for a given update entry (AJAX) ──────────────────────────────

@login_required
def api_extra_tasks(request):
    """Return extra (non-plan) tasks for a given member+date update as JSON."""
    member_pk = (request.GET.get('member') or '').strip()
    date_str = (request.GET.get('date') or '').strip()
    if not member_pk or not date_str:
        return JsonResponse({'ok': False, 'items': []})
    entry_d = _parse_date(date_str)
    if not entry_d:
        return JsonResponse({'ok': False, 'items': []})
    updates = DailyEntry.objects.filter(
        kind='update',
        member_id=member_pk if member_pk.isdigit() else None,
        entry_date=entry_d,
    ).prefetch_related('items')
    items = []
    for upd in updates:
        for it in upd.items.filter(plan_item__isnull=True):
            items.append({
                'id': it.id,
                'entry_id': upd.id,
                'task_type': it.task_type,
                'question_type': it.question_type or '',
                'customer': it.customer or '',
                'count': it.count or '',
                'notes': it.notes or '',
                'status': it.status,
                'due_at': it.due_at.isoformat() if it.due_at else '',
                'jira_key': it.jira_issue_key or '',
                'jira_url': it.jira_issue_url or '',
            })
    return JsonResponse({'ok': True, 'items': items})


# ── XLSX export ──────────────────────────────────────────────────────────────

@login_required
def export_xlsx(request):
    """Export the current filtered work-log as an xlsx file."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    today = date.today()
    week_start = _monday(today)
    week_end = week_start + timedelta(days=6)
    from_d = _parse_date(request.GET.get('from'), default=week_start)
    to_d   = _parse_date(request.GET.get('to'),   default=week_end)
    sel_member = (request.GET.get('member') or '').strip()
    sel_kind   = (request.GET.get('kind') or '').strip().lower()
    sel_status = (request.GET.get('status') or '').strip().lower()
    q          = (request.GET.get('q') or '').strip()

    qs = DailyEntry.objects.select_related('member').prefetch_related('items').filter(
        entry_date__range=(from_d, to_d)
    )
    if sel_member:
        qs = qs.filter(member__display_name=sel_member)
    if sel_kind in ('plan', 'update'):
        qs = qs.filter(kind=sel_kind)
    if sel_status in ('open', 'in_progress', 'blocked', 'closed'):
        qs = qs.filter(status=sel_status)
    if q:
        from django.db.models import Q as DQ
        qs = qs.filter(
            DQ(raw_text__icontains=q) |
            DQ(items__notes__icontains=q) |
            DQ(items__task_type__icontains=q) |
            DQ(items__customer__icontains=q)
        ).distinct()

    rows = []
    for e in qs.order_by('entry_date', 'member__display_name'):
        items = list(e.items.all())
        base = dict(
            date=e.entry_date.isoformat(),
            kind=e.kind,
            status=e.status,
            member=e.member.display_name,
        )
        if not items:
            rows.append({**base, 'task_type': '', 'question_type': '',
                         'customer': '', 'count': '', 'notes': e.raw_text or '',
                         'jira_key': e.jira_issue_key or ''})
        else:
            for it in items:
                rows.append({**base,
                              'task_type': it.task_type or '',
                              'question_type': it.question_type or '',
                              'customer': it.customer or '',
                              'count': it.count or '',
                              'notes': it.notes or e.raw_text or '',
                              'jira_key': it.jira_issue_key or e.jira_issue_key or ''})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Work Log'

    # ── Header row styling
    HDR_FILL = PatternFill('solid', fgColor='0F1620')
    HDR_FONT = Font(name='Calibri', bold=True, color='4FFFB0', size=10)
    CELL_FONT = Font(name='Calibri', size=10)
    BORDER = Border(bottom=Side(style='thin', color='1A2535'))
    ALT_FILL = PatternFill('solid', fgColor='0C1117')

    headers = ['Date', 'Kind', 'Status', 'Member', 'Task Type',
               'Q Type', 'Customer', 'Count', 'Jira', 'Notes']
    col_widths = [12, 8, 12, 20, 28, 16, 18, 7, 14, 50]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = 'A2'

    STATUS_MAP = {'open': 'Open', 'in_progress': 'In Progress',
                  'blocked': 'Blocked', 'closed': 'Done'}

    for ri, r in enumerate(rows, 2):
        vals = [
            r['date'], r['kind'].capitalize(),
            STATUS_MAP.get(r['status'], r['status']),
            r['member'], r['task_type'], r['question_type'],
            r['customer'], r['count'] if r['count'] else '',
            r['jira_key'], r['notes'],
        ]
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = CELL_FONT
            cell.alignment = Alignment(vertical='center', wrap_text=(ci == 10))
            cell.border = BORDER
            if fill:
                cell.fill = fill

    # Auto-filter
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fname = f'content_dashboard_{from_d.isoformat()}_{to_d.isoformat()}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response


# ── members ───────────────────────────────────────────────────────────────────

@login_required
def members_list(request):
    members = Member.objects.filter(is_active=True).annotate(
        entry_count=Count('entries', distinct=True)
    ).order_by('display_name')
    return render(request, 'tracker/members_list.html', {'members': members})


@login_required
def member_detail(request, member_id):
    member = get_object_or_404(Member, id=member_id)
    today = date.today()
    week_start = _monday(today)
    week_end = week_start + timedelta(days=6)
    sel_date = _parse_date(request.GET.get('date'), default=today)

    day_entries = DailyEntry.objects.filter(
        member=member, entry_date=sel_date
    ).prefetch_related('items').order_by('created_at')

    week_entries = DailyEntry.objects.filter(
        member=member, entry_date__range=(week_start, week_end)
    ).prefetch_related('items').order_by('entry_date', 'created_at')

    def flatten(qs):
        out = []
        for e in qs:
            items = list(e.items.all())
            base = dict(
                date=e.entry_date.isoformat(),
                kind=e.kind,
                status=e.status,
                jira_key=e.jira_issue_key or '',
                jira_url=e.jira_issue_url or '',
            )
            if not items:
                out.append({
                    **base,
                    'item_id': None,
                    'task_type': '',
                    'question_type': '',
                    'customer': '',
                    'count': 0,
                    'notes': e.raw_text or '',
                    'status': e.status,
                    'jira_key': e.jira_issue_key or '',
                    'jira_url': e.jira_issue_url or '',
                })
            else:
                for it in items:
                    jk = it.jira_issue_key or e.jira_issue_key or ''
                    ju = it.jira_issue_url or e.jira_issue_url or ''
                    out.append({
                        **base,
                        'item_id': it.id,
                        'task_type': it.task_type,
                        'question_type': it.question_type or '',
                        'customer': it.customer or '',
                        'count': int(it.count or 0),
                        'notes': it.notes or e.raw_text or '',
                        'status': it.status,
                        'jira_key': jk,
                        'jira_url': ju,
                    })
        return out

    day_rows  = flatten(day_entries)
    week_rows = flatten(week_entries)

    totals = {
        'week_items':   len([r for r in week_rows if r['task_type']]),
        'week_volume':  sum(r['count'] for r in week_rows),
        'week_plans':   len([r for r in week_rows if r['kind'] == 'plan']),
        'week_updates': len([r for r in week_rows if r['kind'] == 'update']),
    }

    by_task = {}
    for r in week_rows:
        if not r['task_type']:
            continue
        t = by_task.setdefault(r['task_type'], {'task_type': r['task_type'], 'items': 0, 'volume': 0})
        t['items'] += 1
        t['volume'] += int(r['count'] or 0)
    by_task_rows = sorted(by_task.values(), key=lambda x: (-x['volume'], -x['items']))

    return render(request, 'tracker/member_detail.html', {
        'member': member,
        'day': sel_date.isoformat(),
        'week_start': week_start.isoformat(),
        'week_end': week_end.isoformat(),
        'day_rows': day_rows,
        'week_rows': week_rows[:300],
        'totals': totals,
        'by_task': by_task_rows,
    })


# ── reports ───────────────────────────────────────────────────────────────────

@login_required
def reports(request):
    today = date.today()
    from_d = _parse_date(request.GET.get('from'), default=today.replace(day=1))
    to_d   = _parse_date(request.GET.get('to'),   default=today)
    sel_status = (request.GET.get('status') or '').strip().lower()
    sel_member = (request.GET.get('member') or '').strip()

    qs = DailyEntry.objects.filter(entry_date__range=(from_d, to_d))
    if sel_status in ('open', 'in_progress', 'blocked', 'closed'):
        qs = qs.filter(status=sel_status)
    if sel_member:
        qs = qs.filter(member__display_name=sel_member)

    by_task = (
        EntryItem.objects.filter(entry__in=qs)
        .values('task_type', 'question_type')
        .annotate(items=Count('id'), total_count=Sum('count'))
        .order_by('-total_count', '-items')
    )
    by_member = (
        qs.values('member__display_name', 'status')
        .annotate(entries=Count('id', distinct=True))
        .order_by('member__display_name')
    )
    members = Member.objects.filter(is_active=True).order_by('display_name')

    # CSV export
    if request.GET.get('export') == 'csv':
        full_qs = qs.select_related('member').prefetch_related('items')
        resp = HttpResponse(content_type='text/csv')
        resp['Content-Disposition'] = (
            f'attachment; filename="content_dashboard_{from_d}_{to_d}.csv"'
        )
        w = csv.writer(resp)
        w.writerow(['date', 'kind', 'status', 'member', 'task_type',
                    'question_type', 'customer', 'count', 'notes', 'source',
                    'jira_key', 'jira_url'])
        for e in full_qs:
            items = list(e.items.all())
            jk = e.jira_issue_key or ''
            ju = e.jira_issue_url or ''
            if not items:
                w.writerow([e.entry_date, e.kind, e.status, e.member.display_name,
                             '', '', '', '', e.raw_text or '', e.source, jk, ju])
            else:
                for it in items:
                    ijk = it.jira_issue_key or jk
                    iju = it.jira_issue_url or ju
                    w.writerow([e.entry_date, e.kind, it.status, e.member.display_name,
                                 it.task_type, it.question_type or '', it.customer or '',
                                 it.count or '', it.notes or e.raw_text or '', e.source,
                                 ijk, iju])
        return resp

    return render(request, 'tracker/reports.html', {
        'from': from_d.isoformat(), 'to': to_d.isoformat(),
        'sel_status': sel_status, 'sel_member': sel_member,
        'by_task': by_task, 'by_member': by_member,
        'members': members,
    })


# ── AE weekly updates ─────────────────────────────────────────────────────────

AE_FIELDS = [
    ('setter_enhancements',      'Setter Enhancements'),
    ('he_support_replies',       '#he_support_v2 Replies / Resolutions'),
    ('eng_assessment_replies',   '#engineering_assessment Replies / Resolutions'),
    ('facecode_replies',         '#engineering_facecode Replies / Resolutions'),
    ('data_requests_replies',    '#data_requests Replies / Resolutions'),
    ('redash_queries',           'Redash Queries'),
    ('bug_fixes',                'Bug Fixes'),
    ('deployments',              'Deployments'),
    ('setter_enhancements_count','Setter Enhancements (count)'),
    ('utilities',                'Utilities'),
]


def _is_ae_or_admin(request):
    if request.user.is_staff or request.user.is_superuser:
        return True
    return Member.objects.filter(
        display_name__iexact=request.user.username, role='ae', is_active=True
    ).exists()


@login_required
def ae_daily(request):
    if not _is_ae_or_admin(request):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Access restricted to Application Engineers and admins.")

    today = date.today()
    week_start = _monday(today)

    from_d = _parse_date(request.GET.get('from'), default=week_start)
    to_d   = _parse_date(request.GET.get('to'),   default=today)
    sel_date = _parse_date(request.GET.get('date'), default=today)

    ae_members = list(Member.objects.filter(is_active=True, role='ae').order_by('display_name'))

    # Today's submission forms — one per AE member
    today_entries = {
        e.member.display_name: e
        for e in AEDailyUpdate.objects.filter(
            entry_date=sel_date, member__in=ae_members
        ).select_related('member')
    }
    member_forms = []
    for m in ae_members:
        entry = today_entries.get(m.display_name)
        field_rows = [
            {'field': f, 'label': lbl, 'value': getattr(entry, f, 0) if entry else 0}
            for f, lbl in AE_FIELDS
        ]
        member_forms.append({
            'member': m,
            'entry': entry,
            'field_rows': field_rows,
            'notes': entry.notes if entry else '',
        })

    # Date-range summary table: rows = dates, columns = members × metrics
    range_qs = AEDailyUpdate.objects.filter(
        entry_date__range=(from_d, to_d),
        member__in=ae_members,
    ).select_related('member').order_by('entry_date')

    # Build {date_str: {member_name: entry}}
    range_by_date = {}
    for e in range_qs:
        ds = e.entry_date.isoformat()
        range_by_date.setdefault(ds, {})[e.member.display_name] = e

    # Aggregate totals per member across date range
    totals = {m.display_name: {f: 0 for f, _ in AE_FIELDS} for m in ae_members}
    for e in range_qs:
        for f, _ in AE_FIELDS:
            totals[e.member.display_name][f] += getattr(e, f, 0)

    # Summary rows for totals table: [{label, values: [per_member_total]}]
    summary_rows = [
        {'label': lbl, 'values': [totals[m.display_name][f] for m in ae_members]}
        for f, lbl in AE_FIELDS
    ]

    # Daily log rows for table: [{date, member_values: [{member, entry}]}]
    sorted_dates = sorted(range_by_date.keys())
    daily_rows = []
    for ds in sorted_dates:
        row_entries = range_by_date[ds]
        daily_rows.append({
            'date': ds,
            'member_entries': [
                {'member': m, 'entry': row_entries.get(m.display_name)}
                for m in ae_members
            ],
        })

    return render(request, 'tracker/ae_daily.html', {
        'ae_members': ae_members,
        'today': today.isoformat(),
        'sel_date': sel_date.isoformat(),
        'from': from_d.isoformat(),
        'to': to_d.isoformat(),
        'fields': AE_FIELDS,
        'member_forms': member_forms,
        'summary_rows': summary_rows,
        'daily_rows': daily_rows,
    })


@login_required
@require_POST
def ae_daily_submit(request):
    if not _is_ae_or_admin(request):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Access restricted to Application Engineers and admins.")

    date_str = request.POST.get('entry_date', '').strip()
    entry_date = _parse_date(date_str)
    if not entry_date:
        messages.error(request, 'Invalid date.')
        return redirect('ae_daily')

    member_id = request.POST.get('member', '').strip()
    if not member_id or not member_id.isdigit():
        messages.error(request, 'Please select a member.')
        return redirect('ae_daily')

    try:
        member = Member.objects.get(pk=int(member_id), role='ae')
    except Member.DoesNotExist:
        messages.error(request, 'Invalid member.')
        return redirect('ae_daily')

    obj, _ = AEDailyUpdate.objects.get_or_create(member=member, entry_date=entry_date)
    for field, _ in AE_FIELDS:
        try:
            setattr(obj, field, max(0, int(request.POST.get(field, 0) or 0)))
        except (TypeError, ValueError):
            setattr(obj, field, 0)
    obj.notes = (request.POST.get('notes') or '').strip() or None
    obj.save()
    messages.success(request, f'Daily update saved for {member.display_name} on {entry_date}.')
    return redirect(f"{reverse('ae_daily')}?date={date_str}&from={request.POST.get('from_d', date_str)}&to={request.POST.get('to_d', date_str)}")


@login_required
def ae_daily_export(request):
    if not _is_ae_or_admin(request):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Access restricted to Application Engineers and admins.")

    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    today = date.today()
    from_d = _parse_date(request.GET.get('from'), default=_monday(today))
    to_d   = _parse_date(request.GET.get('to'),   default=today)

    ae_members = list(Member.objects.filter(is_active=True, role='ae').order_by('display_name'))
    range_qs = AEDailyUpdate.objects.filter(
        entry_date__range=(from_d, to_d), member__in=ae_members
    ).select_related('member').order_by('entry_date')

    by_date = {}
    for e in range_qs:
        ds = e.entry_date.isoformat()
        by_date.setdefault(ds, {})[e.member.display_name] = e

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'AE Daily'

    HDR_FILL = PatternFill('solid', fgColor='0F1620')
    HDR_FONT = Font(name='Calibri', bold=True, color='4FFFB0', size=10)
    CELL_FONT = Font(name='Calibri', size=10)
    METRIC_FONT = Font(name='Calibri', size=10, bold=True)
    BORDER = Border(bottom=Side(style='thin', color='1A2535'))
    ALT_FILL = PatternFill('solid', fgColor='0C1117')

    # Columns: Date | metric_member1 | metric_member2 | ... (one col per member per metric)
    # Layout: Date | [Ayush: Setter Enh] | [Ayush: #he_support] | ... | [Chaitanya: Setter Enh] | ...
    headers = ['Date']
    col_widths = [13]
    for m in ae_members:
        for _, lbl in AE_FIELDS:
            headers.append(f"{m.display_name}\n{lbl}")
            col_widths.append(22)

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 40
    ws.freeze_panes = 'B2'

    sorted_dates = sorted(by_date.keys())
    for ri, ds in enumerate(sorted_dates, 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        c = ws.cell(row=ri, column=1, value=ds)
        c.font = CELL_FONT
        c.border = BORDER
        if fill:
            c.fill = fill
        ci = 2
        for m in ae_members:
            entry = by_date[ds].get(m.display_name)
            for field, _ in AE_FIELDS:
                val = getattr(entry, field, 0) if entry else ''
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = CELL_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = BORDER
                if fill:
                    cell.fill = fill
                ci += 1

    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fname = f'ae_daily_{from_d.isoformat()}_{to_d.isoformat()}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response
